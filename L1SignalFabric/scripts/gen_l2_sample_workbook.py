"""Generate `docs/L2_Connector_Sample_Data.xlsx` — 10 human-style sample records
per connector, each run through `l2.record.project_record` so the resulting L2
records (Org / Entity / Ops facets) are *computed*, not asserted, and together
cover every L2 record type in the four L1→L2 contracts.

Run:  python scripts/gen_l2_sample_workbook.py
"""
from __future__ import annotations

import json
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.signal import Lineage, Operation, SignalEvent, SourceSystem
from l2.record import project_record

T = "demo-tenant"
TS = datetime(2026, 6, 9, 8, 0, tzinfo=timezone.utc)

# ----------------------------------------------------------------- styling ----
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUB_FONT = Font(italic=True, size=10, color="595959")
L2_FILL = PatternFill("solid", fgColor="E2EFDA")
L2_FONT = Font(color="375623", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _ev(entity, ss, data, key=None, meta=None) -> SignalEvent:
    return SignalEvent(
        entity=entity, key=key or {"id": data.get("id") or data.get("crew_id") or "k"},
        source_system=ss, tenant_id=T, operation=Operation.DELTA, data=data,
        timestamp=TS, lineage=Lineage(extraction_id="sample"), metadata=meta or {},
    )


def facet_summary(event: SignalEvent) -> str:
    """Human-readable one-cell summary of the L2 records this event produces."""
    rec = project_record(event)
    parts = []
    for f in rec.facets:
        if f.map.value == "org":
            tgt = f.node_id or f.label
            parts.append(f"OrgMap · {f.kind}:{f.label}")
        elif f.map.value == "entity":
            parts.append(f"EntityMap · {f.event}")
        elif f.map.value == "ops":
            an = f" ({f.agent_name})" if f.agent_name else ""
            parts.append(f"OpsMap · {f.event_type}{an}")
    return "  +  ".join(parts)


# =========================================================================== #
#  Sample data per connector                                                  #
# =========================================================================== #
# Each entry: (columns dict for display, SignalEvent built from the same data)

def crewdb_rows():
    # crew_id, pool, name, rank, grade, nationality, vessel, port, joining_date,
    # experience_years, certifications, status, op
    rows = [
        ("SOF-3002", "signoff", "Oleksandr Tkachenko", "Second Engineer", "Grade A", "Ukrainian", "MT Crude Titan", "Rotterdam", "2025-11-11", 12, "", "Onboard", "DELTA"),
        ("CR-8FE3CF", "signoff", "Suthesh Murti", "Bosun", "Grade C", "Indian", "MV Pacific Star", "Singapore", "2025-12-02", 5, "", "Onboard", "DELTA"),
        ("SOF-2000", "signon", "Miguel Torres", "Chief Officer", "Grade A", "Filipino", "MV Pacific Star", "Singapore", "2025-11-26", 11, "", "Signed Off", "DELTA"),
        ("CR-6630B8", "signon", "Will Smith", "Second Officer", "Grade A", "American", "MV Pacific Star", "Singapore", "2026-07-03", 11, "", "Assigned", "DELTA"),
        ("SNO-1000", "signon", "Juan dela Cruz", "Chief Officer", "Grade A", "Filipino", "Available", "Singapore", "", 12, "STCW Basic Safety; GMDSS; Advanced Fire Fighting; Medical First Aid", "Available", "DELTA"),
        ("SNO-1005", "signon", "Dmitri Volkov", "Master", "Grade A", "Russian", "Available", "Singapore", "", 20, "STCW Basic Safety; GMDSS; ECDIS; BRM", "Available", "DELTA"),
        ("SNO-1011", "signon", "Piotr Kowalski", "Chief Officer", "Grade A", "Polish", "Available", "Rotterdam", "", 13, "STCW Basic Safety; GMDSS; Advanced Fire Fighting; ECDIS", "Available", "DELTA"),
        ("SNO-1002", "signon", "Alexei Petrov", "Chief Engineer", "Grade A", "Ukrainian", "Available", "Rotterdam", "", 15, "STCW Basic Safety; High Voltage; ECDIS", "Available", "DELTA"),
        ("SOF-3001", "signon", "Emilio Navarro", "Chief Officer", "Grade A", "Filipino", "MV Pacific Star", "Singapore", "2025-09-27", 10, "", "Signed Off", "DELTA"),
        ("SNO-1015", "signon", "Jose Mendoza", "Deck Cadet", "Grade D", "Filipino", "Available", "Manila", "", 1, "STCW Basic Safety", "Available", "DELETE"),
    ]
    cols = ["crew_id", "pool", "name", "rank", "grade", "nationality", "vessel",
            "port", "joining_date", "experience_years", "certifications", "status", "op"]
    out = []
    for r in rows:
        d = dict(zip(cols, r))
        data = {k: v for k, v in d.items() if k != "op" and v != ""}
        if data.get("certifications"):
            data["certifications"] = [c.strip() for c in data["certifications"].split(";")]
        if "experience_years" in data:
            data["experience_years"] = int(data["experience_years"])
        ev = _ev("crew", SourceSystem.CREW_DB, data, key={"crew_id": d["crew_id"]},
                 meta={"op": d["op"], "schemaVersion": "1.0"})
        out.append((d, ev))
    return cols, out


def contract_rows():
    rows = [
        ("CT-SOF-2000", "SOF-2000", "Miguel Torres", "Chief Officer", "MV Pacific Star", "Singapore", "2025-11-26", "2026-05-26", "Active", "Engagement"),
        ("CT-SOF-3002", "SOF-3002", "Oleksandr Tkachenko", "Second Engineer", "MT Crude Titan", "Rotterdam", "2025-11-11", "2026-05-11", "Active", "Engagement"),
        ("CT-SOF-3004", "SOF-3004", "Andreas Pappas", "Bosun", "MV Mediterranean Queen", "Piraeus", "2025-12-16", "2026-06-16", "Active", "Engagement"),
        ("CT-SOF-3005", "SOF-3005", "Ferdinand Aquino", "Cook", "MV Atlantic Voyager", "Manila", "2025-12-01", "2026-06-01", "Active", "Engagement"),
        ("CT-SOF-2002", "SOF-2002", "Ramesh Nair", "Chief Engineer", "MV Indian Ocean Pride", "Dubai", "2025-10-27", "2026-04-27", "Active", "Engagement"),
        ("CT-SOF-2009", "SOF-2009", "Li Mingyang", "Second Engineer", "MT Crude Titan", "Shanghai", "2025-11-13", "2026-05-13", "Active", "Engagement"),
        ("CT-SOF-2013", "SOF-2013", "Bohdan Kravchenko", "Master", "MT Crude Titan", "Rotterdam", "2025-08-19", "2026-02-19", "Completed", "Engagement"),
        ("CT-SOF-2005", "SOF-2005", "Konstantinos Diakos", "Master", "MV Mediterranean Queen", "Piraeus", "2025-09-16", "2026-03-16", "Active", "Engagement"),
        ("CT-SOF-2011", "SOF-2011", "Andrzej Wisniewski", "Chief Officer", "MV Atlantic Voyager", "Rotterdam", "2025-10-23", "2026-04-23", "Active", "Engagement"),
        ("CT-SOF-2007", "SOF-2007", "Suresh Iyer", "Electrician", "MV Atlantic Voyager", "Mumbai", "2025-12-04", "2026-06-04", "Active", "Engagement"),
    ]
    cols = ["contract_id", "crew_id", "crew_name", "rank", "vessel", "port",
            "start_date", "end_date", "status", "type"]
    out = []
    for r in rows:
        d = dict(zip(cols, r))
        ev = _ev("contract", SourceSystem.CONTRACT_CLM, dict(d),
                 key={"contract_id": d["contract_id"]}, meta={"schemaVersion": "1.0"})
        out.append((d, ev))
    return cols, out


def vesselport_rows():
    rows = [
        ("VS-001", "vessel", "MV Pacific Star", "9456789", "Pacific Fleet", "Panama", "Container", "In Service"),
        ("VS-002", "vessel", "MV Indian Ocean Pride", "9456790", "Pacific Fleet", "Singapore", "Bulk Carrier", "In Service"),
        ("VS-003", "vessel", "MV Atlantic Voyager", "9456791", "Atlantic Fleet", "Liberia", "Container", "In Service"),
        ("VS-004", "vessel", "MT Crude Titan", "9456792", "Tanker Division", "Marshall Is.", "Crude Tanker", "In Service"),
        ("VS-005", "vessel", "MV Mediterranean Queen", "9456793", "Mediterranean Fleet", "Malta", "RoRo", "In Service"),
        ("PT-SIN", "port", "Singapore", "SGSIN", "", "Singapore", "", "Operational"),
        ("PT-RTM", "port", "Rotterdam", "NLRTM", "", "Netherlands", "", "Operational"),
        ("PT-MNL", "port", "Manila", "PHMNL", "", "Philippines", "", "Operational"),
        ("PT-PIR", "port", "Piraeus", "GRPIR", "", "Greece", "", "Operational"),
        ("PT-DXB", "port", "Dubai", "AEDXB", "", "U.A.E.", "", "Operational"),
    ]
    cols = ["id", "type", "name", "imo_or_locode", "fleet", "country", "vessel_type", "status"]
    out = []
    for r in rows:
        d = dict(zip(cols, r))
        data = {k: v for k, v in d.items() if v != ""}
        # mark the name field as vessel/port so the entity facet branches correctly
        if d["type"] == "port":
            data["port"] = d["name"]
        else:
            data["vessel"] = d["name"]
        ev = _ev("vessel_port", SourceSystem.VESSEL_PORT_DB, data,
                 key={"id": d["id"]}, meta={"schemaVersion": "1.0"})
        out.append((d, ev))
    return cols, out


def slack_rows():
    rows = [
        ("message", "crew-changes", "ops.bridget",
         "Sign-off confirmed for Andreas Pappas (Bosun) from MV Mediterranean Queen at Piraeus. Reliever en route.", "wf-3004"),
        ("message", "crew-ops", "ops.bridget",
         "Sign-on: Piotr Kowalski (Chief Officer) joining MV Pacific Star at Singapore on 2026-07-03.", "wf-3001"),
        ("message", "crew-changes", "manning.li",
         "Crew Member: Ferdinand Aquino\nRank: Cook\nVessel: MV Atlantic Voyager\nPort: Manila\nAction: Sign-off", ""),
        ("reaction", "crew-changes", "fleet.dir",
         ":white_check_mark: on the Pacific Star sign-off thread", ""),
        ("channel_join", "fleet-pacific", "newcrew.raj", "", ""),
        ("message", "general", "ops.bridget",
         "Morning all — ETA update: MV Pacific Star alongside Singapore by 0600 local.", ""),
        ("message", "crew-changes", "manning.li",
         "Heads up: medical certificate for crew id CR-8FE3CF (Suthesh Murti) expires next month.", ""),
        ("reaction", "fleet-atlantic", "ops.bridget", ":anchor: voyage plan looks good", ""),
        ("channel_join", "compliance", "auditor.kate", "", ""),
        ("message", "crew-ops", "ops.bridget",
         "Sign-off initiated: Oleksandr Tkachenko (Second Engineer), MT Crude Titan, Rotterdam.", "wf-3002"),
    ]
    cols = ["entity", "channel", "user", "text", "workflow_id"]
    out = []
    for r in rows:
        d = dict(zip(cols, r))
        data = {"user": d["user"], "user_name": d["user"],
                "channel": d["channel"], "channel_name": d["channel"]}
        if d["text"]:
            data["text"] = d["text"]
        meta = {}
        if d["workflow_id"]:
            meta = {"workflow_id": d["workflow_id"], "l2Intent": "CREATE_SIGNOFF_EVENT"}
        ev = _ev(d["entity"], SourceSystem.SLACK, data,
                 key={"channel": d["channel"], "user": d["user"]}, meta=meta)
        out.append((d, ev))
    return cols, out


def email_rows(ss: SourceSystem):
    rows = [
        ("crewing@oceanic.example", "master.pacificstar@oceanic.example",
         "Sign-off: Andreas Pappas (Bosun) — MV Mediterranean Queen, Piraeus", "thr-9001", "CREATE_SIGNOFF_EVENT", "wf-3004"),
        ("crewing@oceanic.example", "agent.singapore@ports.example",
         "Joining arrangements — Piotr Kowalski (Chief Officer), MV Pacific Star", "thr-9002", "", "wf-3001"),
        ("port.agent@rotterdam.example", "crewing@oceanic.example",
         "RE: Berth window and pilot booking — Rotterdam", "thr-9003", "", ""),
        ("travel@oceanic.example", "crewing@oceanic.example",
         "Flight itinerary issued — relief crew to Singapore", "thr-9002", "", ""),
        ("crewing@oceanic.example", "master.crudetitan@oceanic.example",
         "Sign-off: Bohdan Kravchenko (Master) — MT Crude Titan, Rotterdam", "thr-9004", "CREATE_SIGNOFF_EVENT", "wf-3002"),
        ("compliance@oceanic.example", "crewing@oceanic.example",
         "Compliance check passed for candidate SNO-1011", "thr-9002", "", ""),
        ("hr@oceanic.example", "all-masters@oceanic.example",
         "Updated manning scale for tanker fleet — please review", "thr-9005", "", ""),
        ("crewing@oceanic.example", "master.medqueen@oceanic.example",
         "Sign-off: Yannis Georgiou (Bosun) — MV Mediterranean Queen, Piraeus", "thr-9006", "CREATE_SIGNOFF_EVENT", ""),
        ("agent.manila@ports.example", "crewing@oceanic.example",
         "Visa status update — Filipino crew batch July", "thr-9007", "", ""),
        ("crewing@oceanic.example", "master.atlantic@oceanic.example",
         "Sign-off: Suresh Iyer (Electrician) — MV Atlantic Voyager, Mumbai", "thr-9008", "CREATE_SIGNOFF_EVENT", ""),
    ]
    cols = ["from", "to", "subject", "thread_id", "l2Intent", "workflow_id"]
    out = []
    for r in rows:
        d = dict(zip(cols, r))
        data = {"from": d["from"], "to": [d["to"]], "subject": d["subject"],
                "thread_id": d["thread_id"]}
        meta = {}
        if d["l2Intent"]:
            meta["l2Intent"] = d["l2Intent"]
        if d["workflow_id"]:
            meta["workflow_id"] = d["workflow_id"]
        ev = _ev("email", ss, data, key={"message_id": d["thread_id"]}, meta=meta)
        out.append((d, ev))
    return cols, out


def notion_rows():
    rows = [
        ("page", "Crew Change SOP — Standard Operating Procedure", "https://notion.so/crew-sop", "ops.bridget", "2026-06-08"),
        ("page", "MV Pacific Star — Handover Notes", "https://notion.so/pstar-handover", "master.pacificstar", "2026-06-07"),
        ("database", "Crew Roster 2026", "https://notion.so/db/roster-2026", "manning.li", "2026-06-09"),
        ("page", "Sign-on Checklist — Deck Officers", "https://notion.so/signon-deck", "manning.li", "2026-06-05"),
        ("page", "Compliance Matrix — Tanker Fleet", "https://notion.so/compliance-tanker", "auditor.kate", "2026-06-06"),
        ("database", "Port Agents Directory", "https://notion.so/db/port-agents", "ops.bridget", "2026-05-30"),
        ("page", "Travel Policy — Crew Repatriation", "https://notion.so/travel-policy", "travel.team", "2026-05-28"),
        ("page", "MT Crude Titan — Voyage Plan Q3", "https://notion.so/titan-voyage-q3", "master.crudetitan", "2026-06-04"),
        ("database", "Certificate Tracker", "https://notion.so/db/cert-tracker", "auditor.kate", "2026-06-09"),
        ("page", "Onboarding Guide — New Cadets", "https://notion.so/onboarding-cadets", "hr.team", "2026-05-20"),
    ]
    cols = ["entity", "title", "url", "author", "last_edited"]
    out = []
    for r in rows:
        d = dict(zip(cols, r))
        ev = _ev(d["entity"], SourceSystem.NOTION,
                 {"title": d["title"], "url": d["url"], "author": d["author"],
                  "last_edited": d["last_edited"]},
                 key={"page_id": d["url"].rsplit("/", 1)[-1]})
        out.append((d, ev))
    return cols, out


def sharepoint_rows():
    rows = [
        ("drive_item", "Manning_Scale_2026.xlsx", "Crewing", "https://sp.example/crewing/Manning_Scale_2026.xlsx", "manning.li", "2026-06-08"),
        ("drive_item", "Crew Matrix - Pacific Fleet.xlsx", "Fleet-Pacific", "https://sp.example/pac/CrewMatrix.xlsx", "fleet.dir", "2026-06-07"),
        ("list_item", "Sign-off Request #4471", "CrewChangeRequests", "https://sp.example/lists/ccr/4471", "ops.bridget", "2026-06-09"),
        ("drive_item", "Compliance_Audit_Tanker.docx", "Compliance", "https://sp.example/comp/audit.docx", "auditor.kate", "2026-06-06"),
        ("list_item", "Sign-on Request #4472", "CrewChangeRequests", "https://sp.example/lists/ccr/4472", "manning.li", "2026-06-09"),
        ("drive_item", "Port_Rotation_Schedule.xlsx", "Operations", "https://sp.example/ops/portsched.xlsx", "ops.bridget", "2026-06-05"),
        ("drive_item", "Crew_Certificates_Register.xlsx", "Compliance", "https://sp.example/comp/certs.xlsx", "auditor.kate", "2026-06-09"),
        ("list_item", "Travel Booking #TR-2218", "TravelBookings", "https://sp.example/lists/travel/2218", "travel.team", "2026-06-08"),
        ("drive_item", "MV Atlantic Voyager - Crew List.pdf", "Fleet-Atlantic", "https://sp.example/atl/crewlist.pdf", "master.atlantic", "2026-06-04"),
        ("list_item", "Incident Report #IR-0096", "HSEReports", "https://sp.example/lists/hse/0096", "hse.team", "2026-06-02"),
    ]
    cols = ["entity", "name", "site", "url", "modified_by", "modified"]
    out = []
    for r in rows:
        d = dict(zip(cols, r))
        ev = _ev(d["entity"], SourceSystem.SHAREPOINT,
                 {"name": d["name"], "site": d["site"], "url": d["url"],
                  "modified_by": d["modified_by"], "modified": d["modified"]},
                 key={"item_id": d["url"].rsplit("/", 1)[-1]})
        out.append((d, ev))
    return cols, out


def database_rows():
    rows = [
        ("crew_assignments", "INSERT", "1001", "crew_id=SOF-2000; vessel=MV Pacific Star; role=Chief Officer"),
        ("crew_assignments", "UPDATE", "1002", "crew_id=CR-6630B8; vessel=MV Pacific Star; status=confirmed"),
        ("travel_bookings", "INSERT", "TR-2218", "crew_id=SNO-1011; route=WAW-SIN; depart=2026-07-01"),
        ("compliance_checks", "INSERT", "CC-5510", "candidate_id=SNO-1011; score=0.97; result=pass"),
        ("port_calls", "INSERT", "PC-8821", "vessel=MV Pacific Star; port=Singapore; eta=2026-07-03"),
        ("certificates", "UPDATE", "CERT-330", "crew_id=CR-8FE3CF; type=Medical; expiry=2026-07-31"),
        ("crew_assignments", "DELETE", "1003", "crew_id=SNO-1015; reason=candidate withdrawn"),
        ("vessels", "UPDATE", "VS-004", "name=MT Crude Titan; status=In Service; next_drydock=2027-01"),
        ("invoices", "INSERT", "INV-7740", "vendor=Travel Co; amount=2150.00; currency=USD"),
        ("crew_pool", "UPDATE", "SNO-1005", "crew_id=SNO-1005; availability=Available; rank=Master"),
    ]
    cols = ["table", "op", "pk", "changed_fields"]
    out = []
    for r in rows:
        d = dict(zip(cols, r))
        ev = _ev(d["table"], SourceSystem.DATABASE,
                 {"table": d["table"], "pk": d["pk"], "changed_fields": d["changed_fields"]},
                 key={"table": d["table"], "pk": d["pk"]}, meta={"op": d["op"]})
        out.append((d, ev))
    return cols, out


# ----- OpsMap workflow events (direct OpsMap contract payloads) ---------------
OPS_EVENTS = [
    # case wf-3001 — healthy sign-on (sign-off of Juan, sign-on of Piotr)
    ("wf-3001", "workflow_created", "", "2026-06-09T08:00:00Z",
     {"crew_name": "Juan dela Cruz", "rank": "Chief Officer", "vessel": "MV Pacific Star", "crew_id": "SOF-2000"}),
    ("wf-3001", "agent_completed", "Crew Matching Agent", "2026-06-09T08:00:12Z",
     {"candidate_name": "Piotr Kowalski", "candidate_rank": "Chief Officer", "candidate_id": "SNO-1011"}),
    ("wf-3001", "agent_completed", "Travel Agent", "2026-06-09T08:00:15Z", {}),
    ("wf-3001", "agent_completed", "Notification Agent", "2026-06-09T08:00:18Z", {}),
    ("wf-3001", "crew_updated", "", "2026-06-09T08:01:00Z", {}),
    ("wf-3001", "auto_compliance", "", "2026-06-09T08:02:30Z",
     {"candidate_name": "Piotr Kowalski", "compliance_score": 0.97}),
    ("wf-3001", "crew_signed_on", "", "2026-06-09T08:02:45Z",
     {"crew_name": "Piotr Kowalski", "compliance_status": "pass", "compliance_score": 0.97}),
    # case wf-3002 — manual sign-on entry, rejected at compliance
    ("wf-3002", "sign_on_initiated", "", "2026-06-09T09:10:00Z",
     {"candidate_name": "A. Candidate", "candidate_id": "SNO-9999", "vessel": "MT Crude Titan"}),
    ("wf-3002", "agent_completed", "Compliance Agent", "2026-06-09T09:12:00Z",
     {"candidate_id": "SNO-9999", "compliance_score": 0.41}),
    ("wf-3002", "sign_on_rejected", "", "2026-06-09T09:14:00Z",
     {"crew_name": "A. Candidate", "compliance_status": "fail", "compliance_score": 0.41,
      "failures": ["GMDSS certificate expired", "Medical not valid for vessel type"]}),
    # case wf-3003 — workflow failure
    ("wf-3003", "workflow_failed", "", "2026-06-09T10:05:00Z",
     {"error": "Travel booking provider timed out after 3 retries"}),
]

ORG_STRUCTURE = [
    ("Oceanic Shipping Lines", "Pacific Fleet", "MV Pacific Star; MV Indian Ocean Pride"),
    ("Oceanic Shipping Lines", "Atlantic Fleet", "MV Atlantic Voyager"),
    ("Meridian Maritime", "Tanker Division", "MT Crude Titan"),
    ("Meridian Maritime", "Mediterranean Fleet", "MV Mediterranean Queen"),
]

MANNING = [
    ("default (all vessels)", "Master", 1), ("default (all vessels)", "Chief Officer", 1),
    ("default (all vessels)", "Second Officer", 1), ("default (all vessels)", "Third Officer", 1),
    ("default (all vessels)", "Chief Engineer", 1), ("default (all vessels)", "Second Engineer", 1),
    ("default (all vessels)", "Third Engineer", 1), ("default (all vessels)", "Bosun", 1),
    ("default (all vessels)", "AB Seaman", 2), ("default (all vessels)", "Electrician", 1),
    ("default (all vessels)", "Cook", 1),
    ("MT Crude Titan (override)", "Second Engineer", 2), ("MT Crude Titan (override)", "Pumpman", 1),
    ("MT Crude Titan (override)", "AB Seaman", 3),
]


# =========================================================================== #
#  Workbook rendering                                                         #
# =========================================================================== #
def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = WRAP
        cell.border = BORDER


def add_title(ws, title, subtitle):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = SUB_FONT
    ws.row_dimensions[1].height = 20


def connector_sheet(wb, name, title, subtitle, cols, rows, widths):
    ws = wb.create_sheet(name)
    add_title(ws, title, subtitle)
    header_row = 4
    headers = cols + ["→ Resulting L2 record(s)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header(ws, len(headers), row=header_row)
    for ri, (d, ev) in enumerate(rows, start=header_row + 1):
        for ci, col in enumerate(cols, 1):
            v = d[col]
            if isinstance(v, list):
                v = "; ".join(v)
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        l2 = ws.cell(row=ri, column=len(headers), value=facet_summary(ev))
        l2.fill = L2_FILL
        l2.font = L2_FONT
        l2.alignment = WRAP_TOP
        l2.border = BORDER
    for i, w in enumerate(widths + [46], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return ws


def plain_sheet(wb, name, title, subtitle, cols, rows, widths, l2_col=None):
    ws = wb.create_sheet(name)
    add_title(ws, title, subtitle)
    header_row = 4
    for i, h in enumerate(cols, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header(ws, len(cols), row=header_row)
    for ri, row in enumerate(rows, start=header_row + 1):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if l2_col is not None and ci == l2_col:
                cell.fill = L2_FILL
                cell.font = L2_FONT
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)

    # ---- Overview ----
    ov = wb.create_sheet("Overview")
    add_title(ov, "L2 Connector Sample Data",
              "10 human-style records per connector → the unified L2 Record (Org / Entity / Ops facets).")
    lines = [
        "",
        "WHAT THIS WORKBOOK IS",
        "Each connector tab holds 10 realistic source records. The green '→ Resulting L2 record(s)'",
        "column is COMPUTED by running every row through l2.record.project_record — so it shows the",
        "actual L2 records each sample produces, not a hand-written guess.",
        "",
        "THE UNIFIED L2 RECORD (see docs/L1_TO_L2_UNIFIED_RECORD.md)",
        "One SignalEvent → one L2Record envelope (provenance) + a list of FACETS, one per map:",
        "   • OrgMap facet    — node / edge / signoff_event (tribal-knowledge graph)",
        "   • EntityMap facet — crew/contract/vessel/port record to MERGE (the canonical graph)",
        "   • OpsMap facet    — a process-mining event-log row (per crew-change case)",
        "A single record can feed several maps at once (e.g. an ERP crew row → Org node + Entity trigger).",
        "",
        "MAPS & THEIR CONTRACTS",
        "   • EntityMap  — L1_TO_L2_ENTITY_EVENT_TRIGGERS.md  (crew/contract/vessel/port)",
        "   • OpsMap     — L1_TO_L2_INGESTION_CONTRACT.md      (workflow event log)",
        "   • OrgMap     — L1_TO_L2_ORGMAP_CONTRACT.md         (Company→Fleet→Vessel + manning)",
        "                  + l2/orgmap.py tribal-knowledge edges (Slack/e-mail)",
        "",
        "TABS",
        "   ERP – CrewDB / Contract / Vessel-Port .... EntityMap triggers (+ Org nodes)",
        "   Slack / Gmail / Outlook .................. OrgMap edges (+ Entity & Ops when a notice carries them)",
        "   Notion / SharePoint / Database .......... OrgMap document/entity nodes",
        "   OpsMap – Workflow Events ................ OpsMap event log (3 complete cases, every event_type)",
        "   OrgMap – Structure / Manning ........... OrgMap ownership tree + manning scale",
        "   Coverage Matrix ........................ every L2 record type ↔ the sample that produces it",
    ]
    for i, ln in enumerate(lines, start=4):
        c = ov.cell(row=i, column=1, value=ln)
        if ln.isupper() and ln.strip():
            c.font = Font(bold=True, color="1F4E78")
    ov.column_dimensions["A"].width = 105

    # ---- connector tabs ----
    c, r = crewdb_rows()
    connector_sheet(wb, "ERP – CrewDB", "ERP · Crew DB  (SourceSystem = CREW_DB)",
                    "Crew master rows → EntityMap crew triggers + an OrgMap Crew node. Covers upserted / signed_on / signed_off / deleted.",
                    c, r, [11, 9, 20, 16, 9, 12, 20, 12, 12, 9, 34, 12, 8])
    c, r = contract_rows()
    connector_sheet(wb, "ERP – Contract", "ERP · Contract / CLM  (SourceSystem = CONTRACT_CLM)",
                    "Engagement contract rows → EntityMap contract.upserted + an OrgMap Contract node.",
                    c, r, [14, 11, 20, 16, 22, 12, 12, 12, 11, 12])
    c, r = vesselport_rows()
    connector_sheet(wb, "ERP – Vessel-Port", "ERP · Vessel / Port DB  (SourceSystem = VESSEL_PORT_DB)",
                    "Reference rows → EntityMap vessel.upserted / port.upserted. A 'type' of port makes a Port node, else a Vessel.",
                    c, r, [8, 9, 22, 14, 18, 14, 13, 12])
    c, r = slack_rows()
    connector_sheet(wb, "Slack", "Slack  (SourceSystem = SLACK)",
                    "Messages / reactions / joins → OrgMap edges (POSTED_IN, REACTED_IN, MEMBER_OF). A sign-on/off notice also yields an Entity trigger; with a workflow_id it yields an Ops event.",
                    c, r, [12, 14, 13, 50, 11])
    c, r = email_rows(SourceSystem.GMAIL)
    connector_sheet(wb, "Gmail", "Gmail  (SourceSystem = GMAIL, metadata only)",
                    "From/To/Subject → OrgMap EMAILED edge. l2Intent = CREATE_SIGNOFF_EVENT makes a SignOffEvent node; a workflow_id makes an Ops event; a clean 'Name (Rank)' subject yields an Entity trigger.",
                    c, r, [26, 30, 46, 10, 22, 11])
    c, r = email_rows(SourceSystem.OUTLOOK)
    connector_sheet(wb, "Outlook", "Outlook  (SourceSystem = OUTLOOK, metadata only)",
                    "Identical shape to Gmail — Microsoft Graph mail. Same OrgMap / SignOffEvent / Ops projection.",
                    c, r, [26, 30, 46, 10, 22, 11])
    c, r = notion_rows()
    connector_sheet(wb, "Notion", "Notion  (SourceSystem = NOTION)",
                    "Pages / databases → OrgMap document nodes (label = entity type).",
                    c, r, [11, 42, 34, 16, 13])
    c, r = sharepoint_rows()
    connector_sheet(wb, "SharePoint", "SharePoint  (SourceSystem = SHAREPOINT)",
                    "Drive items / list items → OrgMap document nodes.",
                    c, r, [11, 34, 20, 40, 14, 12])
    c, r = database_rows()
    connector_sheet(wb, "Database", "Database  (SourceSystem = DATABASE, generic SQL CDC/outbox)",
                    "Outbox rows → OrgMap entity nodes keyed by (table, pk). op carries INSERT/UPDATE/DELETE.",
                    c, r, [18, 9, 12, 50])

    # ---- OpsMap workflow events ----
    ops_rows = []
    for case_id, et, an, ts, data in OPS_EVENTS:
        # map event_type/agent_name to the activity L2 records (per the contract)
        activity = {
            "workflow_created": "Sign-Off Initiated", "crew_updated": "Sign-Off Confirmed",
            "auto_compliance": "Compliance Check", "sign_on_initiated": "Compliance Check (manual)",
            "crew_signed_on": "Signed On  [TERMINAL ✓]", "sign_on_rejected": "Sign-On Rejected  [TERMINAL ✗]",
            "workflow_failed": "Workflow Failed  [TERMINAL ✗]",
            "agent_completed": {"Crew Matching Agent": "Crew Matching", "Travel Agent": "Travel Arranged",
                                "Notification Agent": "Crew Notified", "Compliance Agent": "Compliance Check"}.get(an, "?"),
        }[et]
        ops_rows.append([case_id, et, an, ts, json.dumps(data), activity])
    plain_sheet(wb, "OpsMap – Workflow Events", "OpsMap · Workflow Event Log  (POST /opsmap/events)",
                "Three complete cases covering every event_type and every agent_name. case_id = workflow_id is the join key; L2 sorts by timestamp and maps each event to the Activity shown.",
                ["case_id", "event_type", "agent_name", "timestamp", "data (JSON)", "→ L2 Activity"],
                ops_rows, [10, 18, 20, 24, 52, 26], l2_col=6)

    # ---- OrgMap structure ----
    plain_sheet(wb, "OrgMap – Structure", "OrgMap · Ownership Structure  (POST /orgmap/structure)",
                "Company → Fleet → Vessel ownership. Vessel names MUST match EntityMap Vessel.name exactly (they are the join key). L2 adds Company/Fleet nodes + OWNS/OPERATES edges over existing vessels.",
                ["company", "fleet", "vessels (must match Vessel.name)", "→ L2 nodes & edges"],
                [[co, fl, ve, f"Company:{co} -OWNS-> Fleet:{fl} -OPERATES-> {ve.count(';')+1} vessel(s)"]
                 for co, fl, ve in ORG_STRUCTURE],
                [26, 22, 40, 44], l2_col=4)

    # ---- OrgMap manning ----
    plain_sheet(wb, "OrgMap – Manning", "OrgMap · Manning Scale  (POST /orgmap/manning)",
                "Required headcount per rank. rank names MUST match crew rank (the join between 'required' and 'have'). 'default' applies to every vessel; a vessel name is a per-ship override. Send 'required' as a number.",
                ["scope", "rank (must match crew rank)", "required", "→ L2 edge"],
                [[sc, rk, rq, f"Vessel/scope -REQUIRES_RANK {{required:{rq}}}-> Rank:{rk}"] for sc, rk, rq in MANNING],
                [26, 30, 10, 50], l2_col=4)

    # ---- Full envelope JSON (one representative record per connector) ----
    reps = [
        ("ERP – CrewDB", "Onboard seafarer entering the sign-off pool", crewdb_rows()[1][0][1]),
        ("ERP – Contract", "Engagement contract for the sign-off", contract_rows()[1][0][1]),
        ("ERP – Vessel-Port", "Vessel reference row", vesselport_rows()[1][0][1]),
        ("Slack", "Sign-on notice carrying a workflow_id", slack_rows()[1][1][1]),
        ("Gmail", "Sign-off e-mail (l2Intent = CREATE_SIGNOFF_EVENT)", email_rows(SourceSystem.GMAIL)[1][0][1]),
        ("Outlook", "Sign-off e-mail via Microsoft Graph", email_rows(SourceSystem.OUTLOOK)[1][0][1]),
        ("Notion", "Knowledge-base page", notion_rows()[1][0][1]),
        ("SharePoint", "Document library file", sharepoint_rows()[1][0][1]),
        ("Database", "Generic CDC/outbox row", database_rows()[1][0][1]),
    ]
    env_rows = []
    for tab, desc, ev in reps:
        rec = project_record(ev)
        env_rows.append([tab, desc,
                         json.dumps(rec.model_dump(mode="json"), indent=2, ensure_ascii=False)])
    es = plain_sheet(wb, "Full Envelope JSON",
                     "Full L2Record Envelope — one representative record per connector",
                     "The complete wire shape L1 emits: provenance header + facets[] fan-out. This is exactly what project_record returns for the named sample.",
                     ["Connector tab", "Representative sample", "Full L2Record (JSON)"],
                     env_rows, [16, 34, 96])
    mono = Font(name="Consolas", size=9)
    for ri in range(5, 5 + len(env_rows)):
        cell = es.cell(row=ri, column=3)
        cell.font = mono
        cell.alignment = WRAP_TOP
        es.row_dimensions[ri].height = 230

    # ---- Coverage matrix ----
    coverage = [
        ("EntityMap", "crew.upserted", "ERP – CrewDB", "SNO-1000 Juan dela Cruz (signon, Available)"),
        ("EntityMap", "crew.signed_on", "ERP – CrewDB", "SOF-2000 Miguel Torres (signon, assigned MV Pacific Star)"),
        ("EntityMap", "crew.signed_off", "ERP – CrewDB", "SOF-3002 Oleksandr Tkachenko (signoff)"),
        ("EntityMap", "crew.deleted", "ERP – CrewDB", "SNO-1015 Jose Mendoza (op = DELETE)"),
        ("EntityMap", "contract.upserted", "ERP – Contract", "CT-SOF-2000 (Miguel Torres engagement)"),
        ("EntityMap", "vessel.upserted", "ERP – Vessel-Port", "VS-001 MV Pacific Star (type = vessel)"),
        ("EntityMap", "port.upserted", "ERP – Vessel-Port", "PT-SIN Singapore (type = port)"),
        ("EntityMap", "crew.signed_off (from notice)", "Slack / Gmail / Outlook", "parsed 'Name (Rank) … vessel at port'"),
        ("OpsMap", "workflow_created", "OpsMap – Workflow", "wf-3001 event 1"),
        ("OpsMap", "agent_completed · Crew Matching Agent", "OpsMap – Workflow", "wf-3001 event 2"),
        ("OpsMap", "agent_completed · Travel Agent", "OpsMap – Workflow", "wf-3001 event 3"),
        ("OpsMap", "agent_completed · Notification Agent", "OpsMap – Workflow", "wf-3001 event 4"),
        ("OpsMap", "agent_completed · Compliance Agent", "OpsMap – Workflow", "wf-3002 event 2"),
        ("OpsMap", "crew_updated", "OpsMap – Workflow", "wf-3001 event 5"),
        ("OpsMap", "auto_compliance", "OpsMap – Workflow", "wf-3001 event 6"),
        ("OpsMap", "sign_on_initiated", "OpsMap – Workflow", "wf-3002 event 1"),
        ("OpsMap", "crew_signed_on (terminal ✓)", "OpsMap – Workflow", "wf-3001 event 7"),
        ("OpsMap", "sign_on_rejected (terminal ✗)", "OpsMap – Workflow", "wf-3002 event 3"),
        ("OpsMap", "workflow_failed (terminal ✗)", "OpsMap – Workflow", "wf-3003 event 1"),
        ("OpsMap", "workflow event from a connector", "Slack / Gmail / Outlook", "notice carrying a workflow_id"),
        ("OrgMap (tribal)", "edge POSTED_IN", "Slack", "message rows"),
        ("OrgMap (tribal)", "edge REACTED_IN", "Slack", "reaction rows"),
        ("OrgMap (tribal)", "edge MEMBER_OF", "Slack", "channel_join rows"),
        ("OrgMap (tribal)", "edge EMAILED", "Gmail / Outlook", "all e-mail rows"),
        ("OrgMap (tribal)", "node SignOffEvent", "Gmail / Outlook", "rows with l2Intent = CREATE_SIGNOFF_EVENT"),
        ("OrgMap (tribal)", "node Crew/Vessel/Port subgraph", "Slack / Gmail", "parsed sign-on/off notices"),
        ("OrgMap (tribal)", "node (Document/Page)", "Notion / SharePoint", "all rows"),
        ("OrgMap (tribal)", "node (generic entity)", "Database", "all outbox rows"),
        ("OrgMap (structure)", "Company / Fleet + OWNS / OPERATES", "OrgMap – Structure", "2 companies, 4 fleets, 5 vessels"),
        ("OrgMap (structure)", "Rank + REQUIRES_RANK {required}", "OrgMap – Manning", "default scale + MT Crude Titan override"),
        ("OrgMap (structure)", "Crew -HAS_RANK-> Rank (derived)", "ERP – CrewDB", "derived by L2 from each crew 'rank'"),
    ]
    plain_sheet(wb, "Coverage Matrix", "Coverage Matrix — every L2 record type ↔ its sample",
                "Proof that the samples exercise every record type across all four contracts.",
                ["L2 Map", "L2 record / facet type", "Produced by tab", "Example sample"],
                [list(x) for x in coverage], [18, 38, 24, 52])

    out = "docs/L2_Connector_Sample_Data.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        out = "docs/L2_Connector_Sample_Data_NEW.xlsx"
        wb.save(out)
        print(f"(canonical file was locked/open — saved to {out} instead)")
    # quick self-check: total facets produced across connector samples
    total = 0
    for fn in (crewdb_rows, contract_rows, vesselport_rows, slack_rows, notion_rows,
               sharepoint_rows, database_rows):
        for _, ev in fn()[1]:
            total += len(project_record(ev).facets)
    for ss in (SourceSystem.GMAIL, SourceSystem.OUTLOOK):
        for _, ev in email_rows(ss)[1]:
            total += len(project_record(ev).facets)
    print(f"saved {out}  ({total} facets produced across connector samples)")


if __name__ == "__main__":
    build()
